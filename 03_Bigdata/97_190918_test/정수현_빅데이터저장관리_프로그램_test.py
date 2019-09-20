import csv
import os
import openpyxl

base_repository_name = 'Bigdata_Repository'
type_folder_1 = 'Type_A'
type_folder_2 = 'Type_B'
file_name = '시뮬레이션_남해군_관광지별_방문객'
dir_delimeter = '/'
file_format = 'csv'
simulation_count = 100
file_size_limit = 10000
simulation_data = ['1111', '상주면', '남해군', '보리암', '1', '14137', '43677']
file_size = 0
number = 1
number_1 = 1
# link = 'C:\\Python_Workspace\\03_Bigdata\\97_190918_test\\'
is_header = False
is_first = True


def getTourPoint_csv(filewriter):
    filewriter.writerow(simulation_data)
    return

def get_dest_file_name(file_index, base_repository_name, file_name, file_format, file_size_limit):
    global is_header
    if file_format == 'csv':
        dest_file_name = f'{base_repository_name}{dir_delimeter}{type_folder_1}{dir_delimeter}{file_name}{str(file_index)}.{file_format}'

        try:
            file_size = os.path.getsize(dest_file_name)
            print(f"'{dest_file_name}' file size: {file_size}")
            print(f"파일당 size 제한: {file_size_limit}")

            if file_size > file_size_limit:
                dest_file_name = f'{base_repository_name}{dir_delimeter}{type_folder_1}{dir_delimeter}{file_name}{str(file_index)+1}.{file_format}'
                is_header = True
            else:
                is_header = False
        except:
            pass
        return dest_file_name
    elif file_format == 'xlsx':
        dest_file_name = f'{base_repository_name}{dir_delimeter}{type_folder_2}{dir_delimeter}{file_name}{str(file_index)}.{file_format}'

        try:
            file_size = os.path.getsize(dest_file_name)
            print(f"'{dest_file_name}' file size: {file_size}")
            print(f"파일당 size 제한: {file_size_limit}")

            if file_size > file_size_limit:
                dest_file_name = f'{base_repository_name}{dir_delimeter}{type_folder_2}{dir_delimeter}{file_name}{str(file_index)+1}.{file_format}'
                is_header = True
            else:
                is_header = False
        except:
            pass
        return dest_file_name

def save_file(file_index, base_repository_name, file_name, file_format, file_size_limit):
    dest_file_name = get_dest_file_name(file_index, base_repository_name, file_name, file_format, file_size_limit)
    global is_header
    global is_first

    if file_format == 'csv':
        csv_out_file = open(dest_file_name, 'a', newline='')
        filewriter = csv.writer(csv_out_file)
        if is_header == True or is_first == True:
            header_list = ['addrCd', 'gungu', 'sido', 'resNm', 'rnum', 'csForCnt', 'csNatCnt']
            filewriter.writerow(header_list)
            is_first = False
            is_header = False

        for index in range(simulation_count):
            getTourPoint_csv(filewriter)
        csv_out_file.close()

    elif file_format == 'xlsx':
        if is_header == True or is_first == True:
            header_list = ['addrCd', 'gungu', 'sido', 'resNm', 'rnum', 'csForCnt', 'csNatCnt']
            if not os.path.exists(dest_file_name):
                exel_output_file = openpyxl.Workbook()
                show_exel = exel_output_file.active
                show_exel.append(header_list)
                exel_output_file.save(dest_file_name)
                is_first = False
                is_header = False
                exel_output_file = openpyxl.load_workbook(filename=dest_file_name, read_only=False, data_only=False)
                add_exel = exel_output_file.active
                for index in range(simulation_count):
                    add_exel.append(simulation_data)
        else:
            exel_output_file = openpyxl.load_workbook(filename=dest_file_name, read_only=False, data_only=False)
            add_exel = exel_output_file.active
            for index in range(simulation_count):
                add_exel.append(simulation_data)

        exel_output_file.save(dest_file_name)

def file_count():
    if file_format == 'csv':
        csv_sub_name = f'{base_repository_name}{dir_delimeter}{type_folder_1}'
        index = len(os.listdir(csv_sub_name))
        return index
    elif file_format == 'xlsx':
        xlsx_sub_name = f'{base_repository_name}{dir_delimeter}{type_folder_2}'
        index = len(os.listdir(xlsx_sub_name))
        return index

while True:
    print("1.환경설정(디렉토리명, 저장 방식...)")
    print("2.작업수행")
    print("3.종료")
    menu = input("메뉴를 선택하세요:")
    if menu == '1':
        dest_file_name = f'{base_repository_name}{dir_delimeter}{file_name}1.{file_format}'
        while True:
            print("1.디렉토리명 초기값: ", base_repository_name)
            print("2.파일명 초기값: ", file_name)
            print("3.포멧 초기값(1.csv 2.xlsx): ", file_format)
            print("4.데이터 용량 제한(byte) 초기값: ", file_size_limit)
            print("5.이전메뉴")
            sub_menu = input("환경설정 메뉴를 선택하세요: ")
            if sub_menu == '1':
                base_repository_name = input("디렉토리명 설정: ")
            elif sub_menu == '2':
                file_name = input("파일명 설정: ")
            elif sub_menu == '3':
                file_format = input("포멧 설정: ")
            elif sub_menu == '4':
                file_size_limit = int(input("용량 설정: "))
            elif sub_menu == '5':
                break
            else:
                print("잘못 입력하셨습니다. 다시 입력하세요.")
                continue

    elif menu == '2':
        dest_file_name = f'{base_repository_name}{dir_delimeter}{file_name}1.{file_format}'
        if file_format == 'csv':
            if number == 1:
                if not os.path.exists(base_repository_name):
                    os.mkdir(base_repository_name)
                if not os.path.exists(f'{base_repository_name}{dir_delimeter}{type_folder_1}'):
                    os.mkdir(f'{base_repository_name}{dir_delimeter}{type_folder_1}')
                    number += 1
                    if not os.path.exists(dest_file_name):
                        save_file(1, base_repository_name, file_name, file_format, file_size_limit)
                    else:
                        save_file(file_count(), base_repository_name, file_name, file_format, file_size_limit)
                else:
                    number += 1
                    dest_file_name = f'{base_repository_name}{dir_delimeter}{type_folder_1}{dir_delimeter}{file_name}1.{file_format}'
                    if not os.path.exists(dest_file_name):
                        save_file(1, base_repository_name, file_name, file_format, file_size_limit)
                    else:
                        save_file(file_count(), base_repository_name, file_name, file_format, file_size_limit)

            elif number != 1:
                dest_file_name = f'{base_repository_name}{dir_delimeter}{type_folder_1}{dir_delimeter}{file_name}1.{file_format}'
                if not os.path.exists(dest_file_name):
                    save_file(1, base_repository_name, file_name, file_format, file_size_limit)
                else:
                    save_file(file_count(), base_repository_name, file_name, file_format, file_size_limit)

        if file_format == 'xlsx':
            if number_1 == 1:
                if not os.path.exists(base_repository_name) and os.path.exists(dest_file_name):
                    os.mkdir(base_repository_name)
                if not os.path.exists(f'{base_repository_name}{dir_delimeter}{type_folder_2}'):
                    os.mkdir(f'{base_repository_name}{dir_delimeter}{type_folder_2}')
                    number_1 += 1
                    if not os.path.exists(dest_file_name):
                        save_file(1, base_repository_name, file_name, file_format, file_size_limit)
                    else:
                        save_file(file_count(), base_repository_name, file_name, file_format, file_size_limit)
                else:
                    number_1 += 1
                    if not os.path.exists(dest_file_name):
                        save_file(1, base_repository_name, file_name, file_format, file_size_limit)
                    else:
                        save_file(file_count(), base_repository_name, file_name, file_format, file_size_limit)
            elif number_1 != 1:
                dest_file_name = f'{file_name}1.{file_format}'
                if not os.path.exists(dest_file_name):
                    save_file(1, base_repository_name, file_name, file_format, file_size_limit)
                else:
                    save_file(file_count(), base_repository_name, file_name, file_format, file_size_limit)
    elif menu == '3':
        break
else:
    print("잘못 입력하셨습니다. 다시 입력하세요.")


